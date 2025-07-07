# src/main.py
import argparse
import logging
import sys
import os
import csv
from datetime import datetime
import time
import threading
from typing import Dict, List
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import uuid
import json # Added for raw JSON output saving

# Add the parent directory (siya/) to the Python path
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

from src.config_loader import load_config
from src.logger import setup_logging
from src.utils import read_companies_from_csv

# Import concrete AI model implementations
from src.ai_models.openai_model import OpenAIModel
from src.ai_models.google_ai_model import GoogleAIModel
from src.ai_models.ollama_model import OllamaModel
from src.ai_models.base import AIBaseModel

# --- Output Configuration ---
OUTPUT_EXCEL_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'data', 'output.xlsx')
OUTPUT_JSON_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'data', 'json_outputs')

# Headers for the "Run Summary" sheet
SUMMARY_HEADERS = [
    "Run ID",
    "Company Name", "Date Time of Run", "AI Model Chosen", "Prompt Used",
    "Error", "Subsidiaries Found Count", "Financial Metrics Found Count",
    "News Items Found Count", "JSON Output File" # Added JSON Output File column
]

# Headers for the "Detailed Extracted Data" sheet
DETAIL_HEADERS = [
    "Run ID",
    "Company Name", "Subsidiary Name", "Subsidiary Location",
    "Financial Metric", "Financial Value", "Financial As Of Date",
    "News Headline", "News Date", "News Summary", "Data Type"
]

def _write_to_excel(data: Dict, output_file_path: str, ai_model_chosen: str, prompt_used: str, error_message: str = None, run_id: str = None, json_output_file: str = None):
    """
    Appends extracted data to an Excel file with two sheets: "Run Summary" and "Detailed Extracted Data".
    Creates the file and sheets with headers if they don't exist.
    """
    siya_logger = logging.getLogger('siya_agent') # Get logger within function scope

    # Ensure the data directory exists
    output_dir = os.path.dirname(output_file_path)
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    workbook = None
    try:
        if os.path.exists(output_file_path):
            workbook = load_workbook(output_file_path)
        else:
            workbook = Workbook()
            if 'Sheet' in workbook.sheetnames:
                workbook.remove(workbook['Sheet'])

        summary_sheet = None
        if "Run Summary" in workbook.sheetnames:
            summary_sheet = workbook["Run Summary"]
        else:
            summary_sheet = workbook.create_sheet("Run Summary")
            summary_sheet.append(SUMMARY_HEADERS)

        detail_sheet = None
        if "Detailed Extracted Data" in workbook.sheetnames:
            detail_sheet = workbook["Detailed Extracted Data"]
        else:
            detail_sheet = workbook.create_sheet("Detailed Extracted Data")
            detail_sheet.append(DETAIL_HEADERS)

        company_name = data.get("company_name", "")
        current_datetime = datetime.now().isoformat()

        subsidiaries_count = len(data.get("subsidiaries", []))
        financial_count = len(data.get("financial_info", []))
        news_count = len(data.get("news_info", []))

        summary_row_data = {
            "Run ID": run_id if run_id else str(uuid.uuid4()),
            "Company Name": company_name,
            "Date Time of Run": current_datetime,
            "AI Model Chosen": ai_model_chosen,
            "Prompt Used": prompt_used,
            "Error": error_message if error_message else "",
            "Subsidiaries Found Count": subsidiaries_count,
            "Financial Metrics Found Count": financial_count,
            "News Items Found Count": news_count,
            "JSON Output File": json_output_file if json_output_file else "" # Add JSON output file path
        }
        summary_sheet.append([summary_row_data.get(header, "") for header in SUMMARY_HEADERS])

        if not error_message:
            subsidiaries = data.get("subsidiaries", [])
            financial_info = data.get("financial_info", [])
            news_info = data.get("news_info", [])

            for sub in subsidiaries:
                detail_row = {
                    "Run ID": run_id if run_id else "",
                    "Company Name": company_name,
                    "Subsidiary Name": sub.get("name", ""),
                    "Subsidiary Location": sub.get("location", ""),
                    "Data Type": "Subsidiary"
                }
                detail_sheet.append([detail_row.get(header, "") for header in DETAIL_HEADERS])

            for fin in financial_info:
                detail_row = {
                    "Run ID": run_id if run_id else "",
                    "Company Name": company_name,
                    "Financial Metric": fin.get("metric", ""),
                    "Financial Value": fin.get("value", ""),
                    "Financial As Of Date": fin.get("as_of_date", ""),
                    "Data Type": "Financial"
                }
                detail_sheet.append([detail_row.get(header, "") for header in DETAIL_HEADERS])

            for news in news_info:
                detail_row = {
                    "Run ID": run_id if run_id else "",
                    "Company Name": company_name,
                    "News Headline": news.get("headline", ""),
                    "News Date": news.get("date", ""),
                    "News Summary": news.get("summary", ""),
                    "Data Type": "News"
                }
                detail_sheet.append([detail_row.get(header, "") for header in DETAIL_HEADERS])
        else:
            siya_logger.warning(f"No detailed data written for '{company_name}' due to processing error.")

        for sheet in [summary_sheet, detail_sheet]:
            for col in sheet.columns:
                max_length = 0
                column = col[0].column
                for cell in col:
                    try:
                        if cell.value is not None:
                            cell_str = str(cell.value)
                            if len(cell_str) > max_length:
                                max_length = len(cell_str)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                if adjusted_width > 100: adjusted_width = 100
                sheet.column_dimensions[get_column_letter(column)].width = adjusted_width

        workbook.save(output_file_path)
        siya_logger.info(f"Data for '{company_name}' saved to '{output_file_path}'.")

    except Exception as e:
        siya_logger.error(f"Error writing to Excel file '{output_file_path}': {e}")
        if workbook:
            try:
                workbook.close()
            except Exception as close_e:
                siya_logger.error(f"Error closing workbook after write error: {close_e}")

def _save_raw_json_output(company_name: str, extracted_data: Dict, ai_model_chosen: str, prompt_used: str, run_id: str, output_dir: str = OUTPUT_JSON_DIR) -> str:
    """
    Saves the raw JSON output from the AI along with metadata to a JSON file.

    Args:
        company_name (str): The name of the company.
        extracted_data (Dict): The raw JSON data extracted by the AI.
        ai_model_chosen (str): The AI model used.
        prompt_used (str): The prompt that was used.
        run_id (str): The unique ID for the current run.
        output_dir (str): Directory to save the JSON files.

    Returns:
        str: The path to the saved JSON file, or None if saving failed.
    """
    siya_logger = logging.getLogger('siya_agent')
    os.makedirs(output_dir, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    # Sanitize company name for filename
    sanitized_company_name = "".join(c for c in company_name if c.isalnum() or c in (' ', '.', '_')).rstrip().replace(' ', '_')
    filename = f"{sanitized_company_name}_{run_id}_{timestamp}.json"
    file_path = os.path.join(output_dir, filename)

    json_output_content = {
        "run_id": run_id,
        "company_name": company_name,
        "ai_model_chosen": ai_model_chosen,
        "prompt_used": prompt_used,
        "timestamp": datetime.now().isoformat(),
        "extracted_data": extracted_data
    }

    try:
        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(json_output_content, f, indent=4, ensure_ascii=False)
        siya_logger.info(f"Raw JSON output saved to: {file_path}")
        return file_path
    except Exception as e:
        siya_logger.error(f"Error saving raw JSON output for '{company_name}' to '{file_path}': {e}")
        return None


def _run_extraction_in_thread(ai_instance, company_name, model_name, prompt_data, result_container):
    """Helper function to run extraction in a separate thread."""
    try:
        # Pass the prompt_data (containing user_template and system_message) to the AI instance's extract_company_info
        extracted_data = ai_instance.extract_company_info(
            company_name,
            model_name,
            prompt_data.get("user_template", ""),
            prompt_data.get("system_message", "")
        )
        result_container['data'] = extracted_data
    except Exception as e:
        result_container['error'] = str(e)
    finally:
        result_container['completed'] = True

def process_single_company(company_name: str, ai_instance: AIBaseModel, model_name: str, config: Dict, siya_logger: logging.Logger, run_id: str, prompt_version: str):
    """
    Processes a single company by calling the AI model and handling timeouts.
    Saves the extracted data (or error) to an Excel file and raw JSON file.
    """
    siya_logger.info(f"Processing company: '{company_name}' with model '{model_name}' using prompt version '{prompt_version}'...")
    extracted_data = {}
    error_message = None
    
    # Get the specific prompt data for the chosen version
    prompt_data = config["PROMPT_TEMPLATES"].get(prompt_version)
    if not prompt_data:
        error_message = f"Prompt version '{prompt_version}' not found in prompts.json. Skipping company."
        siya_logger.error(error_message)
        _write_to_excel({"company_name": company_name}, OUTPUT_EXCEL_PATH, model_name, prompt_version, error_message, run_id, None) # Pass None for json_output_file
        return

    # The prompt_used for logging and Excel summary will be the user_template content
    prompt_used_text = prompt_data.get("user_template", "").replace("[COMPANY_PLACEHOLDER]", company_name)
    
    result_container = {'data': {}, 'error': None, 'completed': False}
    timeout_seconds = config.get("COMPANY_RESEARCH_TIMEOUT_SECONDS", 60)

    # Pass the prompt_data dictionary to the thread for extraction
    thread = threading.Thread(target=_run_extraction_in_thread, args=(ai_instance, company_name, model_name, prompt_data, result_container))
    thread.start()
    thread.join(timeout=timeout_seconds)

    json_output_file_path = None

    if not result_container['completed']:
        error_message = f"Processing for '{company_name}' timed out after {timeout_seconds} seconds."
        siya_logger.error(error_message)
        extracted_data = {"company_name": company_name}
    elif result_container['error']:
        error_message = f"Error during AI extraction for '{company_name}': {result_container['error']}"
        siya_logger.error(error_message)
        extracted_data = {"company_name": company_name}
    else:
        extracted_data = result_container['data']
        if "error" in extracted_data:
            error_message = f"AI model returned an error for '{company_name}': {extracted_data['error']}"
            siya_logger.error(error_message)
        
        json_output_file_path = _save_raw_json_output(company_name, extracted_data, model_name, prompt_used_text, run_id)

    _write_to_excel(extracted_data, OUTPUT_EXCEL_PATH, model_name, prompt_used_text, error_message, run_id, json_output_file_path)

    if error_message:
        siya_logger.warning(f"Skipping '{company_name}' due to error/timeout. Details logged to Excel and console.")
    else:
        siya_logger.info(f"Successfully processed '{company_name}'. Data saved to '{OUTPUT_EXCEL_PATH}' and raw JSON to '{json_output_file_path}'.")


def main():
    """
    Main entry point for the SIYA agent application.
    Handles configuration loading, logging setup, and CLI argument parsing.
    Orchestrates interactive AI provider and model selection, and company processing.
    """
    config = load_config()
    siya_logger = setup_logging(log_level=config.get("LOG_LEVEL", "INFO"))

    siya_logger.info("SIYA Agent started.")
    siya_logger.debug(f"Loaded configuration: {config}")

    parser = argparse.ArgumentParser(
        description="Subsidiary Intelligence Yielding Analyst (SIYA) Agent."
    )
    parser.add_argument(
        "--company",
        type=str,
        help="Name of the single company to research (e.g., 'PepsiCo')."
    )
    parser.add_argument(
        "--csv_file",
        type=str,
        help="Path to a CSV file containing a list of companies."
    )
    parser.add_argument(
        "--provider",
        type=str,
        choices=["openai", "google_ai", "ollama"],
        help="Choose the AI provider to use (openai, google_ai, ollama). Required if not interactive.",
        required=False
    )
    parser.add_argument(
        "--model",
        type=str,
        help="Specific AI model name to use (e.g., 'gpt-4o', 'gemini-pro', 'llama2'). "
             "If not provided in interactive mode, a list of available models will be shown."
    )
    parser.add_argument(
        "--interactive",
        action="store_true",
        help="Run in interactive mode, prompting for choices."
    )
    parser.add_argument( # New argument for prompt version
        "--prompt_version",
        type=str,
        help="Specific prompt version to use (e.g., 'subsidiary_research_v1'). "
             "If not provided in interactive mode, a list of available prompts will be shown."
    )


    args = parser.parse_args()

    current_run_id = str(uuid.uuid4())
    siya_logger.info(f"Starting new run with ID: {current_run_id}")

    ai_instance = None
    selected_prompt_version = None
    
    if args.interactive:
        siya_logger.info("Running in interactive mode.")
        chosen_provider = None

        available_providers = ["openai", "google_ai", "ollama"]
        provider_display_names = {
            "openai": "OpenAI (GPT Models)",
            "google_ai": "Google AI (Gemini Models)",
            "ollama": "Ollama (Local LLMs)"
        }
        preferred_provider = available_providers[0]

        siya_logger.info("\n--- Choose an AI Provider ---")
        for i, provider in enumerate(available_providers):
            display_name = provider_display_names.get(provider, provider)
            highlight_marker = " *" if provider == preferred_provider else ""
            print(f"{i+1}. {display_name}{highlight_marker}")

        while chosen_provider not in available_providers:
            try:
                choice_input = input(f"Enter your choice (1-{len(available_providers)}) or press Enter for default ({preferred_provider}): ").strip()
                if not choice_input:
                    chosen_provider = preferred_provider
                    print(f"Using default provider: {chosen_provider}")
                elif choice_input.isdigit() and 1 <= int(choice_input) <= len(available_providers):
                    chosen_provider = available_providers[int(choice_input) - 1]
                else:
                    print("Invalid choice. Please enter a number from the list or press Enter for default.")
            except Exception as e:
                print(f"An error occurred during selection: {e}")
                siya_logger.error(f"Error during provider selection: {e}")
                return

        args.provider = chosen_provider
        siya_logger.info(f"Selected AI Provider: {args.provider}")

        try:
            if args.provider == "openai":
                ai_instance = OpenAIModel(config)
            elif args.provider == "google_ai":
                ai_instance = GoogleAIModel(config)
            elif args.provider == "ollama":
                ai_instance = OllamaModel(config)
            else:
                siya_logger.error(f"AI provider '{args.provider}' is not yet implemented or recognized.")
                return

            if ai_instance:
                available_models = ai_instance.list_available_models()
                if not available_models:
                    siya_logger.error(f"No models found for {args.provider}. Check API key/base URL and connection.")
                    return

                preferred_model = ai_instance.get_preferred_model()
                if preferred_model not in available_models:
                    siya_logger.warning(f"Preferred model '{preferred_model}' not found among available models. Falling back to first available.")
                    preferred_model = available_models[0]

                siya_logger.info(f"\n--- Choose a Model for {args.provider} ---")
                for i, model in enumerate(available_models):
                    highlight_marker = " *" if model == preferred_model else ""
                    print(f"{i+1}. {model}{highlight_marker}")

                chosen_model = None
                while chosen_model not in available_models:
                    try:
                        model_choice_input = input(f"Enter your choice (1-{len(available_models)}) or press Enter for default ({preferred_model}): ").strip()
                        if not model_choice_input:
                            chosen_model = preferred_model
                            print(f"Using default model: {chosen_model}")
                        elif model_choice_input.isdigit() and 1 <= int(model_choice_input) <= len(available_models):
                            chosen_model = available_models[int(model_choice_input) - 1]
                        else:
                            print("Invalid choice. Please enter a number from the list or press Enter for default.")
                    except Exception as e:
                        print(f"An error occurred during model selection: {e}")
                        siya_logger.error(f"Error during model selection: {e}")
                        return

                args.model = chosen_model
                siya_logger.info(f"Selected AI Model: {args.model}")

        except ValueError as e:
            siya_logger.error(f"Configuration error for {args.provider}: {e}")
            return
        except ConnectionError as e:
            siya_logger.error(f"Connection error for {args.provider}: {e}. Is the server running?")
            return
        except Exception as e:
            siya_logger.error(f"Failed to initialize AI instance for {args.provider}: {e}")
            return
        
        available_prompts = list(config["PROMPT_TEMPLATES"].keys())
        if not available_prompts:
            siya_logger.error("No prompt templates found in config/prompts.json. Exiting.")
            return

        preferred_prompt_version = config.get("DEFAULT_PROMPT_VERSION", available_prompts[0])
        if preferred_prompt_version not in available_prompts:
            siya_logger.warning(f"Default prompt version '{preferred_prompt_version}' not found. Falling back to first available: {available_prompts[0]}.")
            preferred_prompt_version = available_prompts[0]

        siya_logger.info("\n--- Choose a Prompt Template ---")
        for i, prompt_v in enumerate(available_prompts):
            description = config["PROMPT_TEMPLATES"].get(prompt_v, {}).get("description", "No description available.")
            highlight_marker = " *" if prompt_v == preferred_prompt_version else ""
            print(f"{i+1}. {prompt_v} ({description}){highlight_marker}")

        chosen_prompt_version = None
        while chosen_prompt_version not in available_prompts:
            try:
                prompt_choice_input = input(f"Enter your choice (1-{len(available_prompts)}) or press Enter for default ({preferred_prompt_version}): ").strip()
                if not prompt_choice_input:
                    chosen_prompt_version = preferred_prompt_version
                    print(f"Using default prompt: {chosen_prompt_version}")
                elif prompt_choice_input.isdigit() and 1 <= int(prompt_choice_input) <= len(available_prompts):
                    chosen_prompt_version = available_prompts[int(prompt_choice_input) - 1]
                else:
                    print("Invalid choice. Please enter a number from the list or press Enter for default.")
            except Exception as e:
                print(f"An error occurred during prompt selection: {e}")
                siya_logger.error(f"Error during prompt selection: {e}")
                return
        
        selected_prompt_version = chosen_prompt_version


    # --- Non-Interactive Mode Logic ---
    else: # Non-interactive mode
        if not args.provider:
            siya_logger.error("Provider must be specified in non-interactive mode. Exiting.")
            return
        if not args.model:
            siya_logger.error("Model must be specified in non-interactive mode if not interactive. Exiting.")
            return
        
        # Determine prompt version for non-interactive mode
        if args.prompt_version:
            selected_prompt_version = args.prompt_version
            if selected_prompt_version not in config["PROMPT_TEMPLATES"]:
                siya_logger.error(f"Specified prompt version '{selected_prompt_version}' not found in prompts.json. Exiting.")
                return
        else:
            selected_prompt_version = config.get("DEFAULT_PROMPT_VERSION")
            if not selected_prompt_version or selected_prompt_version not in config["PROMPT_TEMPLATES"]:
                siya_logger.error(f"No prompt version specified and DEFAULT_PROMPT_VERSION is not set or found. Exiting.")
                return
        
        siya_logger.info(f"Running non-interactive with Provider: {args.provider}, Model: {args.model}, Prompt: {selected_prompt_version}")

        try:
            if args.provider == "openai":
                ai_instance = OpenAIModel(config)
            elif args.provider == "google_ai":
                ai_instance = GoogleAIModel(config)
            elif args.provider == "ollama":
                ai_instance = OllamaModel(config)
            else:
                siya_logger.error(f"AI provider '{args.provider}' is not yet implemented or recognized for non-interactive mode.")
                return
        except ValueError as e:
            siya_logger.error(f"Configuration error for {args.provider} in non-interactive mode: {e}")
            return
        except ConnectionError as e:
            siya_logger.error(f"Connection error for {args.provider} in non-interactive mode: {e}. Is the server running?")
            return
        except Exception as e:
            siya_logger.error(f"Failed to initialize AI instance for {args.provider} in non-interactive mode: {e}")
            return

    # Ensure a prompt version is selected before proceeding
    if not selected_prompt_version:
        siya_logger.error("No prompt version selected. Exiting.")
        return

    # --- Core Agent Orchestration (Task 11) ---
    if ai_instance and args.company:
        process_single_company(args.company, ai_instance, args.model, config, siya_logger, current_run_id, selected_prompt_version)
    elif ai_instance and args.csv_file:
        siya_logger.info(f"Processing companies from CSV: {args.csv_file}")
        companies_to_process = read_companies_from_csv(args.csv_file)
        if companies_to_process:
            siya_logger.info(f"Starting batch processing of {len(companies_to_process)} companies from '{args.csv_file}'.")
            for company_name in companies_to_process:
                if company_name:
                    process_single_company(company_name, ai_instance, args.model, config, siya_logger, current_run_id, selected_prompt_version)
                else:
                    siya_logger.warning("Skipping empty company name found in CSV.")
        else:
            siya_logger.warning(f"No companies found in CSV file: {args.csv_file}. Please check the file content.")
    else:
        siya_logger.info("No company or CSV file specified. Use --company or --csv_file argument.")

    siya_logger.info("SIYA Agent finished.")

if __name__ == "__main__":
    main()
